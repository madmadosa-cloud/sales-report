# -*- coding: utf-8 -*-
"""
매출 분석 프로그램 — 데스크톱 앱

브라우저 없이 독립 창으로 매출분석.html 을 실행합니다.

실행:  run_app.bat  또는  python app_launcher.py
"""

from __future__ import annotations

import functools
import json
import os
import re
import shutil
import subprocess
import sys
import threading
import time
import urllib.request
from http.server import HTTPServer, SimpleHTTPRequestHandler
from pathlib import Path

SERVER_PORT = 17890


def app_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def session_dir() -> Path:
    return app_dir() / "저장데이터"


def export_dir() -> Path:
    d = app_dir() / "다운로드"
    d.mkdir(parents=True, exist_ok=True)
    return d


def safe_export_filename(name: str, default_ext: str = ".xls") -> str:
    base = Path(name).name
    base = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", base).strip()
    ext = default_ext.lower()
    if not base.lower().endswith(ext):
        if "." in base:
            base = Path(base).stem + ext
        else:
            base += ext
    return base or ("export" + ext)


def sanitize_xlsx_sheet_name(name: str) -> str:
    s = re.sub(r'[\\/*?:\[\]]', " ", str(name)).strip()
    return (s[:31] if s else "Sheet")


def open_folder(path: Path) -> None:
    folder = str(path.resolve())
    if sys.platform == "win32":
        os.startfile(folder)
    elif sys.platform == "darwin":
        subprocess.run(["open", folder], check=False)
    else:
        subprocess.run(["xdg-open", folder], check=False)


def save_excel_export(payload: dict) -> dict:
    filename = safe_export_filename(payload.get("filename") or "export.xls", ".xls")
    content = payload.get("content") or ""
    out_dir = export_dir()
    out_path = out_dir / filename
    out_path.write_text(content, encoding="utf-8-sig")
    open_folder(out_dir)
    return {"ok": True, "path": str(out_path), "folder": str(out_dir)}


def save_excel_html_report(payload: dict) -> dict:
    """보고서용 HTML 엑셀 — 차트 PNG를 _files 폴더에 저장해 연결 깨짐 방지."""
    import base64
    import html as html_mod

    filename = safe_export_filename(payload.get("filename") or "export.xls", ".xls")
    sheets = payload.get("sheets") or []
    if not sheets:
        raise ValueError("보낼 시트가 없습니다.")

    out_dir = export_dir()
    stem = Path(filename).stem
    files_dir = out_dir / f"{stem}_files"
    files_dir.mkdir(parents=True, exist_ok=True)

    def esc(s: object) -> str:
        return html_mod.escape("" if s is None else str(s))

    body: list[str] = []
    img_idx = 0
    for sheet_data in sheets:
        sheet_name = esc(sheet_data.get("name") or "Sheet")
        body.append(
            f"<h2 style='font-size:14pt;margin:18px 0 8px;border-bottom:1px solid #ccc;'>{sheet_name}</h2>"
        )

        for chart in sheet_data.get("charts") or []:
            png_raw = chart.get("b64") or ""
            if not png_raw:
                png_field = chart.get("png") or ""
                if "," in png_field:
                    png_raw = png_field.split(",", 1)[1]
                else:
                    png_raw = png_field
            if not png_raw:
                continue
            img_name = f"chart_{img_idx}.png"
            (files_dir / img_name).write_bytes(base64.b64decode(png_raw))
            rel = f"{files_dir.name}/{img_name}"
            w = int(chart.get("width") or 820)
            h = int(chart.get("height") or 420)
            title = esc(chart.get("title") or "차트")
            body.append(f"<p style='font-weight:bold;margin:6px 0 4px;'>{title}</p>")
            body.append(
                f"<img src='{esc(rel)}' width='{w}' height='{h}' "
                f"style='display:block;margin:0 0 14px;' alt='{title}'/>"
            )
            img_idx += 1

        columns = sheet_data.get("columns") or []
        if columns:
            body.append(
                "<table border='1' cellspacing='0' cellpadding='4' "
                "style='border-collapse:collapse;font-size:11pt;margin-bottom:24px;'>"
            )
            body.append("<tr>")
            for col in columns:
                body.append(
                    f"<th style='background:#eef1f9;font-weight:bold;'>{esc(col)}</th>"
                )
            body.append("</tr>")
            numeric = set(sheet_data.get("numeric") or [])
            for row in sheet_data.get("rows") or []:
                body.append("<tr>")
                for ci, value in enumerate(row):
                    col_name = columns[ci] if ci < len(columns) else ""
                    if col_name in numeric and isinstance(value, (int, float)):
                        style = " style='mso-number-format:\"\\#\\,\\#\\#0\";text-align:right;'"
                        body.append(f"<td{style}>{value}</td>")
                    else:
                        body.append(f"<td>{esc(value)}</td>")
                body.append("</tr>")
            body.append("</table>")

    sheet_title = esc(payload.get("sheetTitle") or "보고서")
    doc = (
        "<!DOCTYPE html><html xmlns:o='urn:schemas-microsoft-com:office:office' "
        "xmlns:x='urn:schemas-microsoft-com:office:excel' "
        "xmlns='http://www.w3.org/TR/REC-html40'>"
        "<head><meta charset='UTF-8'>"
        "<meta name='ProgId' content='Excel.Sheet'>"
        "<!--[if gte mso 9]><xml><x:ExcelWorkbook><x:ExcelWorksheets>"
        f"<x:ExcelWorksheet><x:Name>{sheet_title}</x:Name>"
        "<x:WorksheetOptions><x:DisplayGridlines/></x:WorksheetOptions>"
        "</x:ExcelWorksheet></x:ExcelWorksheets></x:ExcelWorkbook></xml><![endif]-->"
        "</head><body>" + "".join(body) + "</body></html>"
    )

    out_path = out_dir / filename
    out_path.write_text("\ufeff" + doc, encoding="utf-8-sig")
    open_folder(out_dir)
    return {
        "ok": True,
        "path": str(out_path),
        "folder": str(out_dir),
        "filesDir": str(files_dir),
        "html": True,
    }


def save_excel_xlsx_export(payload: dict) -> dict:
    import base64
    import io

    try:
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.styles import Font
    except ImportError as exc:
        raise RuntimeError(
            "차트 포함 엑셀 저장에 openpyxl이 필요합니다. setup_env.py 또는 "
            "pip install openpyxl 을 실행해 주세요."
        ) from exc

    filename = safe_export_filename(payload.get("filename") or "export.xlsx", ".xlsx")
    sheets = payload.get("sheets") or []
    if not sheets:
        raise ValueError("보낼 시트가 없습니다.")

    wb = Workbook()
    wb.remove(wb.active)

    for sheet_data in sheets:
        title = sanitize_xlsx_sheet_name(sheet_data.get("name") or "Sheet")
        ws = wb.create_sheet(title=title)
        row_idx = 1
        numeric = set(sheet_data.get("numeric") or [])
        columns = sheet_data.get("columns") or []

        for chart in sheet_data.get("charts") or []:
            chart_title = chart.get("title") or "차트"
            cell = ws.cell(row=row_idx, column=1, value=chart_title)
            cell.font = Font(bold=True)
            row_idx += 1

            png_raw = chart.get("png") or ""
            if "," in png_raw:
                png_raw = png_raw.split(",", 1)[1]
            if png_raw:
                png_bytes = base64.b64decode(png_raw)
                stream = io.BytesIO(png_bytes)
                stream.seek(0)
                img = XLImage(stream)
                img.width = int(chart.get("width") or 820)
                img.height = int(chart.get("height") or 420)
                ws.add_image(img, f"A{row_idx}")
                row_idx += max(18, int(img.height / 18) + 2)

        if columns:
            for col_idx, col_name in enumerate(columns, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=col_name)
                cell.font = Font(bold=True)
            row_idx += 1

            for row in sheet_data.get("rows") or []:
                for col_idx, value in enumerate(row, start=1):
                    col_name = columns[col_idx - 1] if col_idx - 1 < len(columns) else ""
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    if col_name in numeric and isinstance(value, (int, float)):
                        cell.number_format = "#,##0"
                row_idx += 1

    out_dir = export_dir()
    out_path = out_dir / filename
    wb.save(out_path)
    open_folder(out_dir)
    return {"ok": True, "path": str(out_path), "folder": str(out_dir)}


def load_session() -> dict:
    root = session_dir()
    if not root.exists():
        return {}
    out: dict = {}
    sales_path = root / "sales.csv"
    if sales_path.exists():
        name_path = root / "sales.name"
        name = name_path.read_text(encoding="utf-8").strip() if name_path.exists() else "sales.csv"
        out["sales"] = {"name": name, "text": sales_path.read_text(encoding="utf-8")}
    elif (root / "item.csv").exists():
        item_path = root / "item.csv"
        name_path = root / "item.name"
        name = name_path.read_text(encoding="utf-8").strip() if name_path.exists() else "item.csv"
        out["sales"] = {"name": name, "text": item_path.read_text(encoding="utf-8")}
    for key in ("purchase",):
        csv_path = root / f"{key}.csv"
        if not csv_path.exists():
            continue
        name_path = root / f"{key}.name"
        name = name_path.read_text(encoding="utf-8").strip() if name_path.exists() else f"{key}.csv"
        out[key] = {"name": name, "text": csv_path.read_text(encoding="utf-8")}
    months_path = root / "months.json"
    if months_path.exists():
        try:
            out["months"] = json.loads(months_path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            pass
    for key, filename in (("itemMap", "item_map.json"), ("custMap", "cust_map.json")):
        map_path = root / filename
        if map_path.exists():
            try:
                out[key] = json.loads(map_path.read_text(encoding="utf-8"))
            except json.JSONDecodeError:
                pass
    meta_path = root / "meta.json"
    if meta_path.exists():
        try:
            out["meta"] = json.loads(meta_path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            pass
    return out


def _save_csv_slot(root: Path, key: str, ent: dict | None) -> None:
    """CSV 슬롯 저장 — payload에 없으면 기존 파일 유지, clear 시에만 삭제."""
    csv_path = root / f"{key}.csv"
    name_path = root / f"{key}.name"
    if ent is None:
        return
    if ent.get("clear"):
        if csv_path.exists():
            csv_path.unlink()
        if name_path.exists():
            name_path.unlink()
        return
    if ent.get("text"):
        csv_path.write_text(ent["text"], encoding="utf-8")
        name_path.write_text(ent.get("name") or f"{key}.csv", encoding="utf-8")


def save_session(data: dict) -> None:
    root = session_dir()
    root.mkdir(parents=True, exist_ok=True)
    sales_ent = data.get("sales")
    if sales_ent is None and data.get("item"):
        sales_ent = data.get("item")
    _save_csv_slot(root, "sales", sales_ent)
    _save_csv_slot(root, "purchase", data.get("purchase"))
    for legacy in ("item", "cust"):
        legacy_csv = root / f"{legacy}.csv"
        legacy_name = root / f"{legacy}.name"
        if sales_ent and sales_ent.get("text") and legacy_csv.exists():
            legacy_csv.unlink()
        if sales_ent and sales_ent.get("text") and legacy_name.exists():
            legacy_name.unlink()
    if "months" in data:
        (root / "months.json").write_text(
            json.dumps(data["months"], ensure_ascii=False),
            encoding="utf-8",
        )
    for key, filename in (("itemMap", "item_map.json"), ("custMap", "cust_map.json")):
        if key in data and isinstance(data[key], dict):
            (root / filename).write_text(
                json.dumps(data[key], ensure_ascii=False),
                encoding="utf-8",
            )
    meta = {
        "savedAt": data.get("savedAt") or time.strftime("%Y-%m-%dT%H:%M:%S"),
    }
    (root / "meta.json").write_text(json.dumps(meta, ensure_ascii=False), encoding="utf-8")


class AppHTTPRequestHandler(SimpleHTTPRequestHandler):
    def do_GET(self) -> None:
        if self.path.split("?", 1)[0] == "/api/session":
            self._send_json(load_session())
            return
        super().do_GET()

    def do_POST(self) -> None:
        path = self.path.split("?", 1)[0]
        if path in (
            "/api/session",
            "/api/export-excel",
            "/api/export-excel-xlsx",
            "/api/export-excel-html-report",
        ):
            length = int(self.headers.get("Content-Length", 0))
            raw = self.rfile.read(length) if length > 0 else b"{}"
            try:
                payload = json.loads(raw.decode("utf-8"))
                if path == "/api/session":
                    save_session(payload)
                    self._send_json({"ok": True})
                elif path == "/api/export-excel-xlsx":
                    self._send_json(save_excel_xlsx_export(payload))
                elif path == "/api/export-excel-html-report":
                    self._send_json(save_excel_html_report(payload))
                else:
                    self._send_json(save_excel_export(payload))
            except Exception as exc:
                self._send_json({"ok": False, "error": str(exc)}, status=500)
            return
        self.send_error(405, "Method Not Allowed")

    def _send_json(self, obj: dict, status: int = 200) -> None:
        body = json.dumps(obj, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def log_message(self, format: str, *args) -> None:
        if self.path.startswith("/api/"):
            return
        super().log_message(format, *args)


def resource_path(name: str) -> Path:
    if getattr(sys, "frozen", False):
        bundled = Path(sys._MEIPASS) / name
        if bundled.exists():
            return bundled
    return app_dir() / name


def find_html() -> Path:
    for name in ("sales_app.html", "매출분석.html"):
        p = resource_path(name)
        if p.exists():
            return p
    htmls = sorted(app_dir().glob("*.html"), key=lambda p: p.stat().st_size, reverse=True)
    if htmls:
        return htmls[0]
    raise FileNotFoundError("HTML file not found in " + str(app_dir()))


def start_html_server(directory: Path, port: int = SERVER_PORT) -> None:
    handler = functools.partial(AppHTTPRequestHandler, directory=str(directory))
    HTTPServer(("127.0.0.1", port), handler).serve_forever()


def app_url(html_name: str, port: int = SERVER_PORT) -> str:
    from urllib.parse import quote

    return f"http://127.0.0.1:{port}/{quote(html_name)}"


def wait_server_ready(port: int = SERVER_PORT, timeout: float = 15.0) -> bool:
    url = f"http://127.0.0.1:{port}/"
    deadline = time.time() + timeout
    while time.time() < deadline:
        try:
            urllib.request.urlopen(url, timeout=1)
            return True
        except Exception:
            time.sleep(0.2)
    return False


def main() -> int:
    try:
        html = find_html()
    except FileNotFoundError as e:
        print(f"[ERROR] {e}")
        return 1

    try:
        import webview
    except ImportError:
        print("pywebview 가 필요합니다. setup_env.py 또는 앱빌드.bat(배포용)을 확인하세요.")
        return 1

    if not getattr(sys, "frozen", False):
        try:
            from setup_env import full_setup

            full_setup(quiet=True)
        except ImportError:
            pass

    root = app_dir()
    if getattr(sys, "frozen", False):
        bundled_html = Path(sys._MEIPASS) / html.name
        target_html = root / html.name
        if bundled_html.exists():
            shutil.copy2(bundled_html, target_html)
            html = target_html
    threading.Thread(target=start_html_server, args=(root, SERVER_PORT), daemon=True).start()
    if wait_server_ready(SERVER_PORT):
        open_url = app_url(html.name, SERVER_PORT)
    else:
        print("[경고] 로컬 서버 대기 시간 초과 — file:// 로 열립니다.")
        open_url = html.resolve().as_uri()
    print(f"앱 주소: {open_url}")
    print(f"저장 폴더: {session_dir()}")

    webview.create_window(
        title="매출 분석 프로그램",
        url=open_url,
        width=1180,
        height=880,
        min_size=(820, 620),
        text_select=True,
    )
    webview.start()
    return 0


if __name__ == "__main__":
    sys.exit(main())
