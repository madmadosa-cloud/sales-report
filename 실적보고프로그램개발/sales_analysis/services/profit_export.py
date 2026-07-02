"""이익분석 보고서 Excel/PDF 출력"""

from __future__ import annotations

from io import BytesIO

from django.template.loader import render_to_string
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from sales_analysis.services.profit_aggregation import ProfitReport

HEADER_FILL = PatternFill(start_color="EEF1F9", end_color="EEF1F9", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="FFF7E6", end_color="FFF7E6", fill_type="solid")
NUMBER_FMT = "#,##0"
QTY_FMT = "#,##0.##"


def build_profit_download_filename(report: ProfitReport, ext: str) -> str:
    label = report.period_label.replace("/", "-").replace("\\", "-").replace(":", "-")
    if report.is_welfare:
        prefix = "이익분석_복지부"
    elif report.is_final:
        prefix = "이익분석_최종"
    else:
        prefix = "이익분석"
    return f"{prefix}_{label}.{ext}"


def export_profit_excel(report: ProfitReport) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = (
        "이익분석_복지부"
        if report.is_welfare
        else "이익분석_최종"
        if report.is_final
        else "이익분석"
    )

    ws.append([f"이익분석 보고서 ({report.period_label})"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append(["판매처(거래처분류)", report.axis_label, "수량", "원가(천원)", "이익금(천원)"])

    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL

    for row in report.rows:
        ws.append(
            [
                row.customer_label,
                row.item_label,
                float(row.quantity),
                row.cost_thousand,
                row.amount_thousand,
            ]
        )
        row_idx = ws.max_row
        ws.cell(row_idx, 3).number_format = QTY_FMT
        ws.cell(row_idx, 4).number_format = NUMBER_FMT
        ws.cell(row_idx, 5).number_format = NUMBER_FMT
        if row.row_type == "total":
            for col in range(1, 6):
                ws.cell(row_idx, col).font = Font(bold=True)
                ws.cell(row_idx, col).fill = TOTAL_FILL
        elif row.row_type == "subtotal":
            for col in range(1, 6):
                ws.cell(row_idx, col).font = Font(bold=True)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_profit_pdf(report: ProfitReport) -> bytes:
    from sales_analysis.services.pdf_export import _register_pdf_font, resolve_korean_font_path
    from xhtml2pdf import pisa

    font_path = resolve_korean_font_path()
    font_name = _register_pdf_font(font_path)
    font_url = font_path.as_uri() if font_path else ""

    html = render_to_string(
        "sales_analysis/profit_report_pdf.html",
        {
            "report": report,
            "font_url": font_url,
            "font_name": font_name,
        },
    )

    result = BytesIO()
    status = pisa.CreatePDF(html, dest=result, encoding="utf-8")
    if status.err:
        raise RuntimeError("PDF 생성 중 오류가 발생했습니다.")
    return result.getvalue()
