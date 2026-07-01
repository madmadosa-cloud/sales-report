"""
Excel 보고서 출력
"""

from __future__ import annotations

from decimal import Decimal
from io import BytesIO

from django.db.models import QuerySet
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from sales_analysis.models import SalesRecord
from sales_analysis.services.aggregation import SalesReport, get_unclassified_records

HEADER_FILL = PatternFill(start_color="EEF1F9", end_color="EEF1F9", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="FFF7E6", end_color="FFF7E6", fill_type="solid")
SUBTOTAL_FILL = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
NUMBER_FMT = "#,##0"
QTY_FMT = "#,##0.##"


def _apply_number_format(ws, row_idx: int, col_count: int, first_numeric_col: int = 3) -> None:
    for c in range(first_numeric_col, col_count + 1):
        ws.cell(row_idx, c).number_format = NUMBER_FMT


def _cell_to_excel_values(cell) -> list:
    qty = float(cell.quantity) if cell.quantity is not None else 0
    return [cell.count, qty, cell.amount_thousand]


def _build_headers(report: SalesReport) -> list[str]:
    headers = [
        "판매처(거래처분류)",
        "품목분류",
        "납품실적(총계) 건수",
        "납품실적(총계) 수량",
        "납품실적(총계) 금액(천원)",
    ]
    for block in report.period_blocks:
        headers.extend([f"{block.label} 건수", f"{block.label} 수량", f"{block.label} 금액(천원)"])
    for m in report.months:
        headers.extend([f"{m}월 건수", f"{m}월 수량", f"{m}월 금액(천원)"])
    return headers


def _row_to_excel(report: SalesReport, row) -> list:
    values = [row.customer_label, row.item_label]
    values.extend(_cell_to_excel_values(row.total))
    for block in report.period_blocks:
        bc = row.blocks.get(block.block_id)
        if bc:
            values.extend(_cell_to_excel_values(bc))
        else:
            values.extend([0, 0, 0])
    for m in report.months:
        mc = row.months.get(m)
        if mc:
            values.extend(_cell_to_excel_values(mc))
        else:
            values.extend([0, 0, 0])
    return values


def export_report_excel(report: SalesReport, qs: QuerySet[SalesRecord]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "매출분석"

    ws.append([report.period_label])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append(_build_headers(report))

    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for row in report.rows:
        excel_row = _row_to_excel(report, row)
        ws.append(excel_row)
        r_idx = ws.max_row
        for c in range(3, len(excel_row) + 1):
            col_mod = (c - 3) % 3
            ws.cell(r_idx, c).number_format = QTY_FMT if col_mod == 1 else NUMBER_FMT
        if row.row_type == "total":
            for c in range(1, len(excel_row) + 1):
                ws.cell(r_idx, c).font = Font(bold=True)
                ws.cell(r_idx, c).fill = TOTAL_FILL
        elif row.row_type == "subtotal":
            for c in range(1, len(excel_row) + 1):
                ws.cell(r_idx, c).font = Font(bold=True)
                ws.cell(r_idx, c).fill = SUBTOTAL_FILL

    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = 12 if col > 2 else 16

    unclassified = get_unclassified_records(qs)
    if unclassified and not report.is_simple:
        ws2 = wb.create_sheet("미분류")
        headers = list(unclassified[0].keys())
        ws2.append(headers)
        for cell in ws2[1]:
            cell.font = Font(bold=True)
        for rec in unclassified:
            ws2.append([rec[h] for h in headers])
        for col_name in ("수량", "합계"):
            if col_name in headers:
                col_idx = headers.index(col_name) + 1
                fmt = QTY_FMT if col_name == "수량" else NUMBER_FMT
                for r in range(2, ws2.max_row + 1):
                    ws2.cell(r, col_idx).number_format = fmt

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_download_filename(report: SalesReport, ext: str) -> str:
    from django.utils import timezone

    today = timezone.localdate().strftime("%Y%m%d")
    if report.start_month == 1 and report.end_month == 6:
        half = "상반기"
    elif report.start_month == 7 and report.end_month == 12:
        half = "하반기"
    else:
        half = f"{report.start_month}-{report.end_month}월"
    if report.is_simple:
        half = f"간편_{half}"
    return f"매출분석_{report.year}년_{half}_{today}.{ext}"
