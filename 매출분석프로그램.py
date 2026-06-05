# -*- coding: cp949 -*-
"""
���� �м� ���α׷�
------------------
ERP���� �ٿ�ε��� ���� �Ǹ���Ȳ CSV�� �о �Ʒ� 3������ �ڵ� �����ϰ�,
�ϳ��� ���� ����(���� ��Ʈ)�� �����Ѵ�.

  1) �ŷ�ó �з��� ����   - '�ŷ�ó�� �Ǹ���Ȳ' ������ �ŷ�ó�ڵ� �� �� �����ڷ� �з�
  2) ǰ�� �з��� ����     - 'ǰ�� �Ǹ���Ȳ' ������ ǰ���ڵ� �� �� �����ڷ� �з�
  3) ���� ǰ�� ����/����  - ǰ���ڵ忡 '����'�� ���Եǰ�, ǰ����� �ູ/�������/���׸���/�Ͼ�����
                            �� ǰ�� ��� �з������귣�庰�� ����/�ݾ� ����

����: ����м����α׷�.py �� �����ϸ� â(GUI)�� ������.
"""

import os
import sys
import datetime
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext

# ---------------------------------------------------------------------------
# �з� �⺻ ��Īǥ (��Ī CSV�� ���� ������ ������ �� ���� ���)
# ---------------------------------------------------------------------------
DEFAULT_ITEM_CATEGORY = {
    "a": "�������",
    "b": "ȭ����, ��Ƽ��",
    "c": "�尩 �Ǻ�",
    "d": "����",
    "e": "�繫��ǰ",
    "f": "�μ����˹�",
    "g": "����",
    "h": "��ǰ",
    "i": "������",
    "j": "������",
    "k": "�뿪����",
    "l": "����ũ",
    "m": "��Ʈ�Ϻ�����",
    "n": "���̹ڽ�",
    "o": "����������",
    "p": "����",
    "q": "LED��",
    "r": "������",
    "s": "�����ε�",
    "t": "���ݼ�����ħ",
    "u": "������",
    "z": "����������",
}

DEFAULT_CUSTOMER_CATEGORY = {
    "z": "��Ÿ",
    "e": "��Ÿ���",
    "d": "�������",
    "c": "�������",
    "b": "�������",
    "a": "������ġ��ü",
}

# ���� 2�� ���Ϳ� ����� �귣��(=���� ��û ��� ȸ��)
DIRECT_DELIVERY_BRANDS = ["�ູ", "�������", "���׸���", "�Ͼ���"]

ENCODINGS = ["utf-8-sig", "cp949", "euc-kr", "utf-8"]


# ---------------------------------------------------------------------------
# ��ƿ
# ---------------------------------------------------------------------------
def to_number(value):
    """'30,000' ���� ���ڿ��� ���ڷ�. ��ų� ��ȯ �Ұ��ϸ� 0."""
    if value is None:
        return 0.0
    s = str(value).strip().replace(",", "").replace('"', "")
    if s in ("", "nan", "None", "-"):
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def first_letter(code):
    s = str(code).strip()
    return s[0].lower() if s else ""


def find_brand(name):
    text = str(name)
    for b in DIRECT_DELIVERY_BRANDS:
        if b in text:
            return b
    return ""


def read_csv_dataframe(path, header_row):
    """���ڵ��� �ڵ� �Ǻ��Ͽ� pandas DataFrame���� �д´�."""
    import pandas as pd

    last_error = None
    for enc in ENCODINGS:
        try:
            df = pd.read_csv(
                path,
                header=header_row,
                dtype=str,
                encoding=enc,
                keep_default_na=False,
                engine="python",
            )
            df.columns = [str(c).strip() for c in df.columns]
            return df
        except Exception as e:  # noqa: BLE001
            last_error = e
    raise RuntimeError(f"������ ���� ���߽��ϴ�: {path}\n{last_error}")


def load_mapping(path, default_map):
    """��Ī CSV(�ڵ�,�̸�)�� �о� {�ҹ����ڵ�: �̸�} ��ųʸ���. �����ϸ� �⺻��."""
    if not path or not os.path.isfile(path):
        return dict(default_map)
    try:
        df = read_csv_dataframe(path, header_row=0)
        if df.shape[1] < 2:
            return dict(default_map)
        key_col, val_col = df.columns[0], df.columns[1]
        mapping = {}
        for _, row in df.iterrows():
            key = str(row[key_col]).strip().lower()
            val = str(row[val_col]).strip()
            if key:
                mapping[key] = val
        return mapping if mapping else dict(default_map)
    except Exception:  # noqa: BLE001
        return dict(default_map)


def pick_column(df, candidates):
    """�ĺ� �̸� �� ���� �����ϴ� �÷����� ��ȯ."""
    for c in candidates:
        if c in df.columns:
            return c
    return None


def amount_series(df, log):
    """��������� �� �ݾ� �÷�(���ް��� �켱, ������ �հ�)�� ���� Series��."""
    col = pick_column(df, ["���ް���", "�հ�", "�ݾ�"])
    if col is None:
        raise RuntimeError("�ݾ� �÷�(���ް���/�հ�)�� ã�� �� �����ϴ�.")
    return df[col].map(to_number), col


# ---------------------------------------------------------------------------
# �ٽ� �м�
# ---------------------------------------------------------------------------
def analyze(item_file, customer_file, item_map_file, customer_map_file,
            output_file, log):
    import pandas as pd

    item_map = load_mapping(item_map_file, DEFAULT_ITEM_CATEGORY)
    customer_map = load_mapping(customer_map_file, DEFAULT_CUSTOMER_CATEGORY)

    # ---------------- �ŷ�ó �з��� (�ŷ�ó�� �Ǹ���Ȳ) ----------------
    log("�� �ŷ�ó�� �Ǹ���Ȳ ���� �д� ��...")
    cdf = read_csv_dataframe(customer_file, header_row=1)
    code_col = pick_column(cdf, ["�ŷ�ó�ڵ�"])
    if code_col is None:
        raise RuntimeError("�ŷ�ó�� ���Ͽ��� '�ŷ�ó�ڵ�' �÷��� ã�� ���߽��ϴ�.")
    # �հ�/�Ұ� �� �ڵ尡 ����ִ� �� ����
    cdf = cdf[cdf[code_col].astype(str).str.strip() != ""].copy()
    c_amount, c_amt_col = amount_series(cdf, log)
    cdf["_�����"] = c_amount
    cdf["_����"] = cdf[pick_column(cdf, ["����"])].map(to_number)
    cdf["_�з��ڵ�"] = cdf[code_col].map(first_letter)
    cdf["�ŷ�ó�з���"] = cdf["_�з��ڵ�"].map(
        lambda k: customer_map.get(k, f"�̺з�({k})" if k else "�̺з�")
    )
    cust_summary = (
        cdf.groupby("�ŷ�ó�з���")
        .agg(�����=("_�����", "sum"), �Ǽ�=("_�����", "size"), ����=("_����", "sum"))
        .reset_index()
        .sort_values("�����", ascending=False)
    )
    log(f"  �ŷ�ó �з� {len(cust_summary)}��, ���� �հ� {int(cust_summary['�����'].sum()):,}��")

    # ---------------- ǰ�� �з��� (ǰ�� �Ǹ���Ȳ) ----------------
    log("�� ǰ�� �Ǹ���Ȳ ���� �д� ��...")
    idf = read_csv_dataframe(item_file, header_row=1)
    icode_col = pick_column(idf, ["ǰ���ڵ�"])
    iname_col = pick_column(idf, ["ǰ���(�԰�)", "ǰ���"])
    if icode_col is None:
        raise RuntimeError("ǰ�� ���Ͽ��� 'ǰ���ڵ�' �÷��� ã�� ���߽��ϴ�.")
    idf = idf[idf[icode_col].astype(str).str.strip() != ""].copy()
    i_amount, i_amt_col = amount_series(idf, log)
    idf["_�����"] = i_amount
    idf["_����"] = idf[pick_column(idf, ["����"])].map(to_number)
    idf["_�з��ڵ�"] = idf[icode_col].map(first_letter)
    idf["ǰ��з���"] = idf["_�з��ڵ�"].map(
        lambda k: item_map.get(k, f"�̺з�({k})" if k else "�̺з�")
    )
    item_summary = (
        idf.groupby("ǰ��з���")
        .agg(�����=("_�����", "sum"), �Ǽ�=("_�����", "size"), ����=("_����", "sum"))
        .reset_index()
        .sort_values("�����", ascending=False)
    )
    log(f"  ǰ�� �з� {len(item_summary)}��, ���� �հ� {int(item_summary['�����'].sum()):,}��")

    # ---------------- ���� (ǰ�� ���� ����) ----------------
    log("�� ���� ǰ�� ���� ��...")
    mask_code = idf[icode_col].astype(str).str.contains("����", na=False)
    if iname_col is not None:
        brand_series = idf[iname_col].map(find_brand)
    else:
        brand_series = idf[icode_col].map(lambda _: "")
    mask_brand = brand_series != ""
    direct = idf[mask_code & mask_brand].copy()
    direct["�귣��"] = brand_series[mask_code & mask_brand]

    direct_by_cat = (
        direct.groupby("ǰ��з���")
        .agg(����=("_����", "sum"), �����=("_�����", "sum"), �Ǽ�=("_�����", "size"))
        .reset_index()
        .sort_values("�����", ascending=False)
    )
    direct_by_brand = (
        direct.groupby("�귣��")
        .agg(����=("_����", "sum"), �����=("_�����", "sum"), �Ǽ�=("_�����", "size"))
        .reset_index()
        .sort_values("�����", ascending=False)
    )
    detail_cols = []
    date_col = pick_column(idf, ["����-No.", "����-No"])
    cust_name_col = pick_column(idf, ["�ŷ�ó��"])
    for c in [date_col, icode_col, iname_col, cust_name_col]:
        if c:
            detail_cols.append(c)
    direct_detail = direct[detail_cols + ["�귣��", "ǰ��з���", "_����", "_�����"]].copy()
    direct_detail = direct_detail.rename(columns={"_����": "����", "_�����": "�����"})
    log(f"  ���� ǰ�� {len(direct)}��, ���� {int(direct['_����'].sum()):,}, "
        f"���� {int(direct['_�����'].sum()):,}��")

    # ---------------- �հ��� �߰� ----------------
    def with_total(df, label_col, num_cols):
        total = {c: "" for c in df.columns}
        total[label_col] = "���հ�"
        for c in num_cols:
            total[c] = df[c].sum()
        return pd.concat([df, pd.DataFrame([total])], ignore_index=True)

    cust_summary = with_total(cust_summary, "�ŷ�ó�з���", ["�����", "�Ǽ�", "����"])
    item_summary = with_total(item_summary, "ǰ��з���", ["�����", "�Ǽ�", "����"])
    direct_by_cat = with_total(direct_by_cat, "ǰ��з���", ["����", "�����", "�Ǽ�"])
    direct_by_brand = with_total(direct_by_brand, "�귣��", ["����", "�����", "�Ǽ�"])

    # ---------------- ���� ���� ----------------
    log("�� ���� ���� �ۼ� ��...")
    sheets = [
        ("�ŷ�ó�з���", cust_summary),
        ("ǰ��з���", item_summary),
        ("����_ǰ��з���", direct_by_cat),
        ("����_�귣�庰", direct_by_brand),
        ("����_��", direct_detail),
    ]
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        for name, df in sheets:
            df.to_excel(writer, sheet_name=name, index=False)
        _format_workbook(writer.book)

    log(f"\n�Ϸ�! ��� ����: {output_file}")
    return output_file


def _format_workbook(wb):
    """��� ����, õ���� �޸�, ���ʺ�, Ʋ����."""
    from openpyxl.styles import Font, PatternFill, Alignment

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    num_keywords = ("�����", "����", "�Ǽ�", "�ݾ�")

    for ws in wb.worksheets:
        # ��� ��Ÿ��
        headers = {}
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            headers[cell.column] = str(cell.value or "")
        # ���� ���� + ���ʺ�
        for col_idx, header in headers.items():
            is_num = any(k in header for k in num_keywords)
            max_len = len(header)
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                if is_num and isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0"
                v = cell.value
                if v is not None:
                    # �ѱ��� ���� �а� ��� ���� ����ġ
                    length = sum(2 if ord(ch) > 127 else 1 for ch in str(v))
                    max_len = max(max_len, length)
            letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 55)
        ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# GUI
# ---------------------------------------------------------------------------
class App:
    def __init__(self, root):
        self.root = root
        root.title("���� �м� ���α׷�")
        root.geometry("760x560")
        root.minsize(680, 500)

        self.item_file = tk.StringVar()
        self.customer_file = tk.StringVar()
        self.item_map_file = tk.StringVar()
        self.customer_map_file = tk.StringVar()
        self.output_file = tk.StringVar()

        self._build_ui()
        self._prefill_defaults()

    # --- UI ���� ---
    def _build_ui(self):
        pad = {"padx": 8, "pady": 4}
        frm = ttk.Frame(self.root, padding=12)
        frm.pack(fill="both", expand=True)

        title = ttk.Label(frm, text="���� �Ǹ���Ȳ ���� �м�",
                          font=("���� ����", 14, "bold"))
        title.grid(row=0, column=0, columnspan=3, sticky="w", pady=(0, 8))

        self._file_row(frm, 1, "�� ǰ�� �Ǹ���Ȳ ����", self.item_file, "csv")
        self._file_row(frm, 2, "�� �ŷ�ó�� �Ǹ���Ȳ ����", self.customer_file, "csv")

        sep = ttk.Separator(frm, orient="horizontal")
        sep.grid(row=3, column=0, columnspan=3, sticky="ew", pady=8)
        opt = ttk.Label(frm, text="�Ʒ� ��Īǥ�� ���û����Դϴ� (����θ� ����� �⺻ �з�ǥ ���)",
                        foreground="#666")
        opt.grid(row=4, column=0, columnspan=3, sticky="w")

        self._file_row(frm, 5, "ǰ��з� ��Ī (����)", self.item_map_file, "csv")
        self._file_row(frm, 6, "�ŷ�ó�з� ��Ī (����)", self.customer_map_file, "csv")

        sep2 = ttk.Separator(frm, orient="horizontal")
        sep2.grid(row=7, column=0, columnspan=3, sticky="ew", pady=8)
        self._file_row(frm, 8, "�� ��� ���� ��ġ", self.output_file, "save")

        run = ttk.Button(frm, text="�м� ����", command=self.run)
        run.grid(row=9, column=0, columnspan=3, sticky="ew", pady=10)

        self.log_box = scrolledtext.ScrolledText(frm, height=12, wrap="word",
                                                 font=("Consolas", 9))
        self.log_box.grid(row=10, column=0, columnspan=3, sticky="nsew")

        frm.columnconfigure(1, weight=1)
        frm.rowconfigure(10, weight=1)

    def _file_row(self, parent, row, label, var, mode):
        ttk.Label(parent, text=label).grid(row=row, column=0, sticky="w", padx=4, pady=4)
        ttk.Entry(parent, textvariable=var).grid(row=row, column=1, sticky="ew", padx=4, pady=4)
        ttk.Button(parent, text="ã�ƺ���...",
                  command=lambda: self._browse(var, mode)).grid(row=row, column=2, padx=4, pady=4)

    def _browse(self, var, mode):
        if mode == "save":
            path = filedialog.asksaveasfilename(
                title="��� ���� ����",
                defaultextension=".xlsx",
                filetypes=[("Excel ����", "*.xlsx")],
            )
        else:
            path = filedialog.askopenfilename(
                title="���� ����",
                filetypes=[("CSV ����", "*.csv"), ("��� ����", "*.*")],
            )
        if path:
            var.set(path)

    def _prefill_defaults(self):
        """���� ����/����ȭ�鿡�� �˷��� ���ϸ��� �ڵ����� ä���ش�."""
        here = os.path.dirname(os.path.abspath(__file__))
        desktop = os.path.dirname(here)  # ���� ����ȭ��
        candidates = {
            self.item_file: [("ǰ��", "�Ǹ�")],
            self.customer_file: [("�ŷ�ó��", "�Ǹ�")],
            self.item_map_file: [("ǰ��з�",)],
            self.customer_map_file: [("�ŷ�ó�з�",)],
        }
        for folder in (here, desktop):
            if not os.path.isdir(folder):
                continue
            try:
                files = os.listdir(folder)
            except OSError:
                continue
            for var, keysets in candidates.items():
                if var.get():
                    continue
                for fname in files:
                    if not fname.lower().endswith(".csv"):
                        continue
                    for keys in keysets:
                        if all(k in fname for k in keys):
                            var.set(os.path.join(folder, fname))
                            break
        # ��� ���� �⺻ ���
        if not self.output_file.get():
            stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            self.output_file.set(os.path.join(here, f"����м����_{stamp}.xlsx"))

    # --- ���� ---
    def log(self, msg):
        self.log_box.insert("end", str(msg) + "\n")
        self.log_box.see("end")
        self.root.update_idletasks()

    def run(self):
        try:
            import pandas  # noqa: F401
            import openpyxl  # noqa: F401
        except ImportError:
            messagebox.showerror(
                "�ʼ� ���̺귯�� ����",
                "pandas / openpyxl �� ��ġ�Ǿ� ���� �ʽ��ϴ�.\n\n"
                "���� ������Ʈ���� �Ʒ��� �����ϼ���:\n"
                "    pip install pandas openpyxl\n\n"
                "�Ǵ� �Բ� ����ִ� '�����ϱ�.bat' �� �����ϸ� �ڵ� ��ġ�˴ϴ�.",
            )
            return

        item = self.item_file.get().strip()
        cust = self.customer_file.get().strip()
        out = self.output_file.get().strip()
        if not item or not os.path.isfile(item):
            messagebox.showwarning("Ȯ��", "�� ǰ�� �Ǹ���Ȳ ������ �����ϼ���.")
            return
        if not cust or not os.path.isfile(cust):
            messagebox.showwarning("Ȯ��", "�� �ŷ�ó�� �Ǹ���Ȳ ������ �����ϼ���.")
            return
        if not out:
            messagebox.showwarning("Ȯ��", "�� ��� ���� ��ġ�� �����ϼ���.")
            return

        self.log_box.delete("1.0", "end")
        self.log("�м��� �����մϴ�...\n")
        try:
            result = analyze(
                item_file=item,
                customer_file=cust,
                item_map_file=self.item_map_file.get().strip() or None,
                customer_map_file=self.customer_map_file.get().strip() or None,
                output_file=out,
                log=self.log,
            )
        except Exception as e:  # noqa: BLE001
            self.log(f"\n[����] {e}")
            messagebox.showerror("����", f"�м� �� ������ �߻��߽��ϴ�:\n\n{e}")
            return

        if messagebox.askyesno("�Ϸ�", "�м��� �������ϴ�. ��� ������ ���� �����?"):
            try:
                os.startfile(result)  # Windows ����
            except Exception:  # noqa: BLE001
                messagebox.showinfo("�ȳ�", f"��� ���� ��ġ:\n{result}")


def main():
    root = tk.Tk()
    App(root)
    root.mainloop()


if __name__ == "__main__":
    main()
